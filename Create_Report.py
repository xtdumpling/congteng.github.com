#!/usr/bin/python3
# Python 3.9.0

# pip install openpyxl

# Debian Linux:
# sudo apt install python3-pip
# pip3 install openpyxl
# pip3 install docx2txt

'''

INPUT: 
SRS_for_test.docx  :  
requirement number: MF-SCU-XXXX
MF-SCU-4112


fortest.xlsm: S203L_FUN_TCXXXX
column J(requirement number): MF-SCU-XXXX
column F(test case number): S203L_FUN_TCXXXX

Output:
create: 
coverage_report.xlsx  list：column B: SRS#   column C: Covered  column D: Test Case#

If exist in fortest.xlsm, covered

'''

import os
import re
import docx2txt
from openpyxl import load_workbook
from openpyxl import Workbook

requirement_pattern = re.compile(r'MF-SCU-\d\d\d\d')
test_case_pattern = re.compile(r'S203L_FUN_TC\d\d\d\d')

os.chdir('/home/xt/Desktop/script')
work_book= load_workbook(keep_vba=True, filename='fortest.xlsx')
sheet = work_book['Sheet1'] # get sheet by name
print(sheet.title)

# Column F is the Test Case Number
column_F = [item.value for item in list(sheet.columns)[ord('F') - ord('A')]]
# Column J is the Requirement Number
column_J = [item.value for item in list(sheet.columns)[ord('J') - ord('A')]]

print(column_F)

print(column_J)

print('ok')

for each_requirement in column_J:
    if len(each_requirement) == 1 and requirement_pattern.match(each_requirement):
        print('Requirement: {}'.format(each_requirement))
    else:
        print(each_requirement)
print('\n')


for each_test_case in column_F:
    if len(each_test_case) == 1 and test_case_pattern.match(each_test_case):
        print('Testcase: {}'.format(each_test_case))
print('\n')

my_text=docx2txt.process('SRS_for_test.docx')
all_requirements = requirement_pattern.findall(my_text)
if len(all_requirements) == 0:
    print('No requirement is found!')


newWorkbook= Workbook()
sheet = newWorkbook.active
print(sheet.title)
newWorkbook.save('Report.xlsx')

newWorkbook = load_workbook('Report.xlsx')
sheet = newWorkbook.create_sheet('Report', 0)
newWorkbook.save('Report.xlsx')
